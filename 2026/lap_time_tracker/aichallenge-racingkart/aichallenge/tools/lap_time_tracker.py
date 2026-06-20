#!/usr/bin/env python3
"""
走行結果を output/lap_time_results.xlsx へ追記するツール。

列構造:
  A列: ファイル名ラベル（config.yaml / reference.launch.xml / 空）
  B列: グループラベル（control / pp / mpc / lqr / map など）
  C列: パラメータ名（"Parameter" ヘッダー）
  D列〜: 各走行の値（右に追記）

使い方:
    python3 lap_time_tracker.py
    python3 lap_time_tracker.py --result /path/to/result-details.json
    python3 lap_time_tracker.py --launch /path/to/reference.launch.xml
    python3 lap_time_tracker.py --config /path/to/config.yaml
    python3 lap_time_tracker.py --section-times '{"A": 12.3, "B": 14.5}'
    python3 lap_time_tracker.py --output /path/to/lap_time_results.xlsx
"""

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import yaml
    _YAML_AVAILABLE = True
except ImportError:
    _YAML_AVAILABLE = False

# ============================================================
# パス設定
# ============================================================
_SCRIPT_DIR = Path(__file__).resolve().parent
_REPO_ROOT = _SCRIPT_DIR.parents[1]  # aichallenge-racingkart/

def _latest_run_d1() -> "Path | None":
    """output/ 直下の最新日付フォルダ内の d1/ を返す（latest を除く）。"""
    output_dir = _REPO_ROOT / "output"
    if not output_dir.exists():
        return None
    folders = [p for p in output_dir.iterdir() if p.is_dir() and p.name != "latest"]
    if not folders:
        return None
    return max(folders, key=lambda p: p.stat().st_mtime) / "d1"

def _default_result_candidates():
    cands = []
    d1 = _latest_run_d1()
    if d1:
        cands.append(d1 / "d1-result-details.json")
        cands.append(d1 / "result-details.json")
    cands.append(_REPO_ROOT / "output" / "latest" / "d1" / "result-details.json")
    return cands

_DEFAULT_SECTION_CANDIDATES = [
    _REPO_ROOT / "output" / "result-section.json",
]

_DEFAULT_LAUNCH_CANDIDATES = [
    _REPO_ROOT / "aichallenge" / "workspace" / "src" / "aichallenge_submit"
    / "aichallenge_submit_launch" / "launch" / "reference.launch.xml",
]

_DEFAULT_CONFIG_CANDIDATES = [
    _REPO_ROOT / "aichallenge" / "workspace" / "src" / "aichallenge_submit"
    / "multi_purpose_mpc_ros" / "config" / "config.yaml",
]

_DEFAULT_OUTPUT = _REPO_ROOT / "output" / "lap_time_results.xlsx"

# ============================================================
# 列定義（A=ファイル名、B=グループ、C=パラメータ名、D〜=値）
# ============================================================
COL_FILE   = 1  # A列: ファイル名ラベル
COL_GROUP  = 2  # B列: グループラベル
COL_PARAM  = 3  # C列: パラメータ名
COL_DATA   = 4  # D列〜: 走行データ

ROW_TITLE = 1

# ============================================================
# スタイル定義
# ============================================================
HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
NO_FILL = PatternFill(fill_type=None)
RESULT_FILLS = [NO_FILL, NO_FILL]
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True, shrink_to_fit=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True, shrink_to_fit=True)

# ============================================================
# ユーティリティ
# ============================================================

def latest_output_folder_name() -> "str | None":
    output_dir = _REPO_ROOT / "output"
    folders = [
        p for p in output_dir.iterdir()
        if p.is_dir() and p.name not in ("latest",)
    ]
    if not folders:
        return None
    return max(folders, key=lambda p: p.stat().st_mtime).name


def find_file(candidates: list) -> "Path | None":
    for p in candidates:
        if p.exists():
            return p
    return None


def load_result(path: Path) -> dict:
    with open(path) as f:
        return json.load(f)


def parse_launch_xml(path: Path) -> dict:
    """reference.launch.xml から arg/let の name と default/value を抽出する。"""
    params = {}
    text = path.read_text()
    for tag in ("arg", "let"):
        attrs = ("default", "value") if tag == "arg" else ("value",)
        for attr in attrs:
            pattern = rf'<{tag}\s[^>]*name="([^"]+)"[^>]*{attr}="([^"]+)"'
            for m in re.finditer(pattern, text):
                name, val = m.group(1), m.group(2)
                if name not in params:
                    params[name] = val
    m = re.search(r'<param name="csv_path"\s+value="([^"]+)"', text)
    if m:
        params["csv_path"] = m.group(1)
    return params


def _flatten_yaml(d: dict, prefix: str = "") -> dict:
    """ネストされた yaml dict をフラット化する（キーはドット区切り）。"""
    out = {}
    for k, v in d.items():
        key = f"{prefix}.{k}" if prefix else k
        if isinstance(v, dict):
            out.update(_flatten_yaml(v, key))
        else:
            out[key] = v
    return out


def parse_config_yaml(path: Path) -> dict:
    """config.yaml からパラメータを抽出する（フラット化済み辞書を返す）。"""
    if not _YAML_AVAILABLE:
        print("[WARN] PyYAML が未インストールのため config.yaml を読み込めません。pip install pyyaml")
        return {}
    with open(path) as f:
        data = yaml.safe_load(f)
    if not isinstance(data, dict):
        return {}
    return _flatten_yaml(data)


def get_csv_basename(params: dict) -> str:
    raw = params.get("csv_path", "")
    return Path(raw).name if raw else ""


# ============================================================
# xlsx 操作
# ============================================================

COL_WIDTH_DATA   = 16.2   # 3cm ≈ 16.2 文字幅
ROW_HEIGHT_TITLE = 28.35  # 1cm ≈ 28.35pt


def ensure_structure(ws):
    """新規xlsx作成時にヘッダー行（1行目）だけ書き込む。行ラベルはExcelで管理する。"""
    for col, val, align in [
        (COL_FILE,  "file",      CENTER),
        (COL_GROUP, "group",     CENTER),
        (COL_PARAM, "Parameter", LEFT),
    ]:
        c = ws.cell(row=ROW_TITLE, column=col)
        c.value = val
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = align
        c.border = THIN_BORDER

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 36


def next_run_col(ws) -> int:
    col = COL_DATA
    while ws.cell(row=ROW_TITLE, column=col).value is not None:
        col += 1
    return col


def find_param_row(ws, param_name: str) -> "int | None":
    """C列（COL_PARAM）でパラメータ名を検索して行番号を返す。"""
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=COL_PARAM).value == param_name:
            return r
    return None


def write_run(ws, result: dict, launch_params: dict, config_params: dict,
              section_times: dict, run_label: str, na_lap: bool = False):
    col = next_run_col(ws)
    run_index = col - COL_DATA
    fill = RESULT_FILLS[run_index % 2]

    def write(row, value):
        c = ws.cell(row=row, column=col)
        c.value = value
        c.alignment = CENTER
        c.border = THIN_BORDER
        c.fill = fill

    # タイトル行
    c = ws.cell(row=ROW_TITLE, column=col)
    c.value = run_label
    c.alignment = CENTER
    c.border = THIN_BORDER
    c.fill = HEADER_FILL
    c.font = HEADER_FONT

    # C列を走査してすべての行に値を書き込む
    laps = result.get("laps", [])
    lap_count  = result.get("lap_count", len(laps))
    total_time = result.get("total_lap_time", None)
    best_lap   = result.get("min_lap_time",   None)
    lap_index  = 0  # lap_N_time の何番目か

    # パラメータ（C列に書かれているラベルを走査して、A列のファイル名に応じて参照元を切り替える）
    _PATH_PARAMS  = {"csv_path_accel_map", "csv_path_brake_map"}
    _SKIP_LABELS  = {"Parameter"}

    for r in range(1, ws.max_row + 1):
        cell_c = ws.cell(row=r, column=COL_PARAM).value
        if not cell_c or not isinstance(cell_c, str):
            continue
        if cell_c in _SKIP_LABELS:
            continue

        pname = cell_c

        # ラップ集計・trajectory の固定ラベル
        if pname == "trajectory_csv":
            write(r, get_csv_basename(launch_params))
            continue
        if pname == "total_laps":
            write(r, "n/a" if na_lap else (lap_count if lap_count else None))
            continue
        if pname == "total_time":
            write(r, "n/a" if na_lap else (f"{total_time:.3f}" if total_time else None))
            continue
        if pname == "best_lap_time":
            write(r, "n/a" if na_lap else (f"{best_lap:.3f}" if best_lap else None))
            continue
        if pname.startswith("lap_") and pname.endswith("_time"):
            if na_lap:
                write(r, "n/a")
            else:
                val = f"{laps[lap_index]:.3f}" if lap_index < len(laps) else None
                write(r, val)
            lap_index += 1
            continue
        if pname.startswith("section_") and pname.endswith("_time"):
            continue  # セクションタイムは後段で処理

        # A列のファイル名ラベルで参照元を判定。
        # A列が未設定でも C列キーが config_params にあれば config_params を優先する。
        file_label = (ws.cell(row=r, column=COL_FILE).value or "").strip()
        if file_label == "config.yaml" or (not file_label and pname in config_params):
            val = config_params.get(pname, None)
        else:
            val = launch_params.get(pname, None)

        if val is None:
            write(r, "null")
        else:
            if pname in _PATH_PARAMS:
                val = Path(val).name
            elif isinstance(val, list):
                val = str(val)
            elif isinstance(val, bool):
                val = str(val).upper()
            write(r, val)

    # セクションタイム（C列の section_X_time で行を検索）
    if not section_times:
        for r in range(1, ws.max_row + 1):
            v = ws.cell(row=r, column=COL_PARAM).value
            if v and str(v).startswith("section_") and str(v).endswith("_time"):
                write(r, "n/a")

    for label, t in section_times.items():
        row = find_param_row(ws, f"section_{label}_time")
        if row is None:
            row = ws.max_row + 1
            for col_i, val_i, align_i in [
                (COL_FILE,  None,                    CENTER),
                (COL_GROUP, None,                    CENTER),
                (COL_PARAM, f"section_{label}_time", LEFT),
            ]:
                c2 = ws.cell(row=row, column=col_i)
                c2.value = val_i
                c2.fill  = NO_FILL
                c2.font  = Font(bold=True, size=10)
                c2.alignment = align_i
                c2.border = THIN_BORDER
        write(row, f"{t:.3f}" if isinstance(t, (int, float)) else ("---" if t is None else t))

    # D列以降の列幅を全列3cm固定、1行目の行高さを1cm固定
    for c in range(COL_DATA, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(c)].width = COL_WIDTH_DATA
    ws.row_dimensions[ROW_TITLE].height = ROW_HEIGHT_TITLE


# ============================================================
# メイン
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="走行結果を xlsx に追記する")
    parser.add_argument("--result", type=Path, default=None,
                        help="result-details.json のパス")
    parser.add_argument("--launch", type=Path, default=None,
                        help="reference.launch.xml のパス")
    parser.add_argument("--config", type=Path, default=None,
                        help="config.yaml のパス")
    parser.add_argument("--output", type=Path, default=_DEFAULT_OUTPUT,
                        help=f"出力 xlsx パス (default: {_DEFAULT_OUTPUT})")
    parser.add_argument("--label", type=str, default=None,
                        help="実行列のラベル（省略時は Run_<timestamp>）")
    parser.add_argument("--section-json", type=Path, default=None,
                        help="result-section.json のパス（省略時は output/ を自動検索）")
    parser.add_argument("--na-lap", action="store_true",
                        help="ラップタイムを n/a として記録する（スタック・手動停止時）")
    args = parser.parse_args()

    # --- result-details.json ---
    result_path = args.result or find_file(_default_result_candidates())
    if args.na_lap:
        result = {}
        if result_path and result_path.exists():
            try:
                result = load_result(result_path)
                print(f"[INFO] result (partial): {result_path}")
            except Exception:
                print("[WARN] result-details.json の読み込みに失敗。ラップタイムは n/a で転記します。")
        else:
            print("[INFO] result-details.json なし。ラップタイムは n/a で転記します。")
    else:
        if result_path is None or not result_path.exists():
            print("[ERROR] result-details.json が見つかりません。--result で指定してください。",
                  file=sys.stderr)
            sys.exit(1)
        print(f"[INFO] result: {result_path}")
        result = load_result(result_path)

    # --- reference.launch.xml ---
    launch_path = args.launch or find_file(_DEFAULT_LAUNCH_CANDIDATES)
    if launch_path is None or not launch_path.exists():
        print("[WARN] reference.launch.xml が見つかりません。パラメータは空欄になります。",
              file=sys.stderr)
        launch_params = {}
    else:
        print(f"[INFO] launch:  {launch_path}")
        launch_params = parse_launch_xml(launch_path)

    # --- config.yaml ---
    config_path = args.config or find_file(_DEFAULT_CONFIG_CANDIDATES)
    if config_path is None or not config_path.exists():
        print("[WARN] config.yaml が見つかりません。config パラメータは空欄になります。",
              file=sys.stderr)
        config_params = {}
    else:
        print(f"[INFO] config:  {config_path}")
        config_params = parse_config_yaml(config_path)

    # --- セクションタイム（result-section.json を自動検索）---
    section_times = {}
    section_path = args.section_json or find_file(_DEFAULT_SECTION_CANDIDATES)
    if section_path and section_path.exists():
        print(f"[INFO] section: {section_path}")
        with open(section_path) as f:
            section_data = json.load(f)
        section_times = dict(section_data.get("section_times", {}))
    else:
        print("[INFO] result-section.json なし。セクションタイムは空欄になります。")

    # --- xlsx を開く or 新規作成 ---
    output_path: Path = args.output
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if output_path.exists():
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        print(f"[INFO] 既存 xlsx に追記: {output_path}")
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "lap_time_results"
        ensure_structure(ws)
        print(f"[INFO] 新規 xlsx を作成: {output_path}")

    run_label = args.label or latest_output_folder_name() or f"Run_{datetime.now().strftime('%Y%m%d-%H%M%S')}"
    write_run(ws, result, launch_params, config_params, section_times, run_label, na_lap=args.na_lap)
    wb.save(output_path)

    print(f"[OK] 書き込み完了: {output_path}  (列: {run_label})")

    laps = result.get("laps", [])
    if args.na_lap:
        print("\n  周回数      : n/a (スタック停止)")
        print("  トータル時間: n/a")
        print("  ベストラップ: n/a")
        for i in range(6):
            print(f"  Lap {i+1}       : n/a")
    else:
        print(f"\n  周回数      : {result.get('lap_count', len(laps))}")
        if result.get("total_lap_time"):
            print(f"  トータル時間: {result['total_lap_time']:.3f} s")
        if result.get("min_lap_time"):
            print(f"  ベストラップ: {result['min_lap_time']:.3f} s")
        for i, t in enumerate(laps[:6]):
            print(f"  Lap {i+1}       : {t:.3f} s")
    if get_csv_basename(launch_params):
        print(f"  Trajectory  : {get_csv_basename(launch_params)}")
    if section_times:
        for label, t in section_times.items():
            print(f"  Section {label}   : {t:.3f} s" if isinstance(t, (int, float)) else f"  Section {label}   : {t}")
    else:
        print("  Section     : (なし)")


if __name__ == "__main__":
    main()
